from fastapi import FastAPI, HTTPException, Query, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import Response
from pydantic import BaseModel, Field

import io
import math
import numpy as np
import pandas as pd
import yfinance as yf
import requests

app = FastAPI(title="Quant Backend")

# --- CORS: keep it simple & permissive ---
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],       # allow all
    allow_credentials=False,   # must be False when origins is "*"
    allow_methods=["*"],
    allow_headers=["*"],
)

# --- Shared HTTP session with UA for yfinance ---
SESSION = requests.Session()
SESSION.headers.update({
    "User-Agent": (
        "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 "
        "(KHTML, like Gecko) Chrome/123.0 Safari/537.36"
    )
})

@app.get("/")
def root():
    return {"message": "API is live"}

@app.get("/health")
def health():
    return {"status": "ok"}

# Simple placeholder assistant
@app.post("/api/assistant")
async def assistant(request: Request):
    payload = await request.json()
    msg = payload.get("message", "")
    return {"text": f"(Echo) You said: {msg}"}

class SummaryIn(BaseModel):
    ticker1: str = Field(..., examples=["NVDA"])
    ticker2: str = Field(..., examples=["AMD"])
    start_date: str = Field(..., examples=["2023-01-01"])
    end_date: str = Field(..., examples=["2024-01-01"])
    initial_invest: float = 1000.0

def _clean_json(x):
    if isinstance(x, float):
        return None if (math.isnan(x) or math.isinf(x)) else float(x)
    if isinstance(x, (np.floating, np.integer)):
        v = x.item()
        if isinstance(v, float) and (math.isnan(v) or math.isinf(v)):
            return None
        return v
    if isinstance(x, (list, tuple)):
        return [_clean_json(v) for v in x]
    if isinstance(x, dict):
        return {k: _clean_json(v) for k, v in x.items()}
    if isinstance(x, np.ndarray):
        return _clean_json(x.tolist())
    return x

def _pick_price_column(df: pd.DataFrame) -> pd.Series:
    if df is None or df.empty:
        raise ValueError("No data returned")
    if isinstance(df, pd.Series):
        return df.astype(float)
    cols = list(df.columns)
    for c in ["Adj Close", "Close"]:
        if c in cols:
            return df[c].astype(float)
    # fallback to first numeric col
    return df.select_dtypes(include=["number"]).iloc[:, 0].astype(float)

def _yf_multi(tickers, start, end) -> pd.DataFrame:
    df = yf.download(
        tickers=tickers,
        start=start,
        end=end,
        auto_adjust=True,
        actions=False,
        interval="1d",
        progress=False,
        repair=True,
        session=SESSION,
        # timeout is supported indirectly via requests
        threads=False,
    )
    if df is None or df.empty:
        return pd.DataFrame()
    if isinstance(df.columns, pd.MultiIndex):
        lvl0 = df.columns.get_level_values(0)
        root = "Adj Close" if "Adj Close" in lvl0 else ("Close" if "Close" in lvl0 else lvl0[0])
        prices = df.loc[:, (root, slice(None))]
        prices.columns = prices.columns.droplevel(0)
        return prices
    # single-level cols (rare with multi)
    return df

def _yf_single(ticker, start, end) -> pd.Series:
    # 1) download
    df = yf.download(
        tickers=ticker,
        start=start,
        end=end,
        auto_adjust=True,
        actions=False,
        interval="1d",
        progress=False,
        repair=True,
        session=SESSION,
        threads=False,
    )
    if df is not None and not df.empty:
        return _pick_price_column(df).rename(ticker)

    # 2) Ticker().history fallback
    th = yf.Ticker(ticker, session=SESSION).history(
        start=start,
        end=end,
        auto_adjust=True,
        actions=False,
        interval="1d",
    )
    if th is not None and not th.empty:
        return _pick_price_column(th).rename(ticker)

    raise ValueError(f"No data returned for {ticker}")

def fetch_prices(t1: str, t2: str, start: str, end: str) -> pd.DataFrame:
    # Try multi first
    try:
        multi = _yf_multi([t1, t2], start, end)
        if not multi.empty:
            keep = [c for c in multi.columns if c in [t1, t2]]
            multi = multi[keep].dropna()
            if multi.shape[0] >= 5 and set([t1, t2]).issubset(multi.columns):
                return multi.astype(float)
    except Exception:
        pass

    # Fallback: fetch individually
    s1 = _yf_single(t1, start, end)
    s2 = _yf_single(t2, start, end)
    prices = pd.concat([s1, s2], axis=1).dropna()
    if prices.shape[0] < 5:
        raise ValueError(
            f"Not enough historical rows after fetching {t1},{t2} "
            f"({prices.shape[0]} rows). Try widening the date range."
        )
    return prices.astype(float)

def compute_summary(t1: str, t2: str, start: str, end: str, capital: float):
    prices = fetch_prices(t1, t2, start, end)
    dates = prices.index.strftime("%Y-%m-%d").tolist()
    s1 = [float(v) for v in prices[t1].tolist()]
    s2 = [float(v) for v in prices[t2].tolist()]

    rets = prices.pct_change().dropna()
    spread = rets[t1] - rets[t2]
    if spread.std(ddof=0) == 0 or np.isnan(spread.std(ddof=0)):
        zscores = [0.0] * len(spread)
    else:
        zscores = ((spread - spread.mean()) / spread.std()).fillna(0).astype(float).tolist()
    z_dates = rets.index.strftime("%Y-%m-%d").tolist()

    joint = {"x": [], "y": [], "z": [], "corr": 0.0}
    try:
        x_grid = np.linspace(rets[t1].min(), rets[t1].max(), 50)
        y_grid = np.linspace(rets[t2].min(), rets[t2].max(), 50)
        X, Y = np.meshgrid(x_grid, y_grid)
        mu = np.array([rets[t1].mean(), rets[t2].mean()], dtype=float)
        cov = np.array(rets[[t1, t2]].cov().values, dtype=float)
        if np.isfinite(cov).all():
            cov = cov + 1e-12 * np.eye(2)
            inv = np.linalg.inv(cov)
            det = float(np.linalg.det(cov))
            diff = np.dstack((X, Y)) - mu
            expo = -0.5 * np.einsum("...i,ij,...j", diff, inv, diff)
            norm = 1.0 / (2.0 * np.pi * np.sqrt(max(det, 1e-24)))
            Z = norm * np.exp(expo)
            corr = float(np.corrcoef(rets[t1].fillna(0), rets[t2].fillna(0))[0, 1])
            joint = {
                "x": x_grid.astype(float).tolist(),
                "y": y_grid.astype(float).tolist(),
                "z": Z.astype(float).tolist(),
                "corr": corr,
            }
    except Exception:
        pass

    windows = [30, 60, 90]
    z_corr = []
    for w in windows:
        r = rets[t1].rolling(w).corr(rets[t2]).to_numpy(dtype=float)
        z_corr.append(r)
    z = np.vstack(z_corr) if z_corr else np.empty((0, 0))
    z_list = [[(None if (v is None or not np.isfinite(v)) else float(v)) for v in row] for row in z.tolist()]
    rolling_surface = {"x_index": list(range(len(rets))), "windows": windows, "z": z_list}

    cum = (1 + rets).cumprod() - 1
    last1 = float(cum[t1].iloc[-1])
    last2 = float(cum[t2].iloc[-1])
    if last1 >= last2:
        outperformer, underperformer = t1, t2
        long_profit = capital * last1
        short_profit = -capital * last2
    else:
        outperformer, underperformer = t2, t1
        long_profit = capital * last2
        short_profit = -capital * last1
    net = float(long_profit + short_profit)
    risk = float(rets[[t1, t2]].std().mean() * np.sqrt(252))

    result = {
        "price_plot": {"dates": dates, "series1": s1, "series2": s2, "ticker1": t1, "ticker2": t2},
        "zscore_plot": {"dates": z_dates, "zscores": zscores},
        "joint_3d": joint,
        "rolling_corr_surface": rolling_surface,
        "backtest": [
            {"Metric": "Outperformer", "Value": outperformer},
            {"Metric": "Underperformer", "Value": underperformer},
            {"Metric": "Net Profit", "Value": f"{net:.2f}"},
            {"Metric": "Annualized Volatility", "Value": f"{risk*100:.2f}%"},
        ],
    }
    return _clean_json(result)

def _excel_rows_from_data(data: dict, t1: str, t2: str, start: str, end: str, capital: float):
    from dateutil.relativedelta import relativedelta

    s1 = pd.Series(data["price_plot"]["series1"], dtype=float)
    s2 = pd.Series(data["price_plot"]["series2"], dtype=float)
    r1 = (float(s1.iloc[-1]) / float(s1.iloc[0]) - 1.0) if len(s1) >= 2 else 0.0
    r2 = (float(s2.iloc[-1]) / float(s2.iloc[0]) - 1.0) if len(s2) >= 2 else 0.0

    if r1 >= r2:
        out_name, out_pct = t1, r1 * 100
        under_name, under_pct = t2, r2 * 100
    else:
        out_name, out_pct = t2, r2 * 100
        under_name, under_pct = t1, r1 * 100

    try:
        risk_str = next(x["Value"] for x in data["backtest"] if x["Metric"] == "Annualized Volatility")
    except StopIteration:
        rets = pd.DataFrame({"A": s1.pct_change(), "B": s2.pct_change()})
        risk = float(rets.std().mean() * np.sqrt(252)) * 100.0
        risk_str = f"{risk:.2f}%"

    try:
        net_profit_val = float(next(x["Value"] for x in data["backtest"] if x["Metric"] == "Net Profit"))
    except Exception:
        net_profit_val = capital * (r1 - r2)

    d1 = pd.to_datetime(start)
    d2 = pd.to_datetime(end)
    rd = relativedelta(d2, d1)
    if rd.years > 0 and rd.months > 0:
        period = f"{rd.years} year{'s' if rd.years>1 else ''}, {rd.months} month{'s' if rd.months>1 else ''}"
    elif rd.years > 0:
        period = f"{rd.years} year{'s' if rd.years>1 else ''}"
    else:
        period = f"{rd.months} month{'s' if rd.months>1 else ''}"

    header = ["Metric", "Value"]
    body = [
        ["Stock Pair", f"{t1} vs {t2}"],
        ["Outperforming Asset", f"{out_name} ({out_pct:.2f}%)"],
        ["Underperforming Asset", f"{under_name} ({under_pct:.2f}%)"],
        ["Net Profit", f"Long ${capital:.0f} {out_name}, Short ${capital:.0f} {under_name}: ${net_profit_val:.2f}"],
        ["Annualized Volatility", f"{risk_str} (Risk Measure)"],
        ["Time Period", period],
    ]
    return [header] + body

def build_excel(data: dict, t1: str, t2: str, start: str, end: str, capital: float) -> bytes:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
        from openpyxl.cell import MergedCell
        HAS_XL = True
    except Exception:
        HAS_XL = False

    rows = _excel_rows_from_data(data, t1, t2, start, end, capital)

    if not HAS_XL:
        df = pd.DataFrame(rows[1:], columns=rows[0])
        return df.to_csv(index=False).encode("utf-8")

    header, body = rows[0], rows[1:]

    wb = Workbook()
    ws = wb.active
    ws.title = "Backtest Report"

    title_text = f"Pairs Trading Report ({t1} vs {t2})"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(header))
    title_cell = ws.cell(row=1, column=1, value=title_text)
    title_cell.font = Font(bold=True, size=14, color="FFFFFF", name="Calibri")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.fill = PatternFill(start_color="00C9A7", end_color="00C9A7", fill_type="solid")

    header_font = Font(bold=True, color="FFFFFF", size=12, name="Calibri")
    header_fill = PatternFill(start_color="00C9A7", end_color="00C9A7", fill_type="solid")
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx, col_name in enumerate(header, 1):
        cell = ws.cell(row=2, column=col_idx, value=col_name)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = header_fill
        cell.border = border

    r = 3
    for row in body:
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r, column=c_idx, value=value)
            cell.alignment = Alignment(wrap_text=True, vertical="center")
            cell.font = Font(size=11, name="Calibri")
            cell.border = border
        r += 1

    for col in ws.columns:
        cells = [cell for cell in col if not isinstance(cell, MergedCell)]
        if not cells:
            continue
        max_len = max(len(str(cell.value)) if cell.value else 0 for cell in cells)
        ws.column_dimensions[cells[0].column_letter].width = max_len + 4

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream.read()

# --- API endpoints ---
@app.post("/api/summary")
def api_summary(body: SummaryIn):
    try:
        return compute_summary(
            body.ticker1.upper(), body.ticker2.upper(),
            body.start_date, body.end_date, body.initial_invest
        )
    except ValueError as e:
        raise HTTPException(status_code=400, detail=f"{e}")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"summary_error: {e}")

@app.get("/api/summary")
def api_summary_get(
    ticker1: str = Query(...),
    ticker2: str = Query(...),
    start_date: str = Query(...),
    end_date: str = Query(...),
    initial_invest: float = Query(1000.0),
):
    try:
        return compute_summary(ticker1.upper(), ticker2.upper(), start_date, end_date, initial_invest)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=f"{e}")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"summary_error: {e}")

@app.post("/api/excel")
def api_excel(body: SummaryIn):
    try:
        data = compute_summary(
            body.ticker1.upper(), body.ticker2.upper(),
            body.start_date, body.end_date, body.initial_invest
        )
        content = build_excel(
            data,
            body.ticker1.upper(), body.ticker2.upper(),
            body.start_date, body.end_date, body.initial_invest
        )
        mime = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        if content[:6] == b"Metric" or content[:5] == b"Metric":
            mime = "text/csv"
        ext = "xlsx" if mime != "text/csv" else "csv"
        filename = f"backtest_{body.ticker1.upper()}_{body.ticker2.upper()}.{ext}"
        return Response(
            content=content,
            media_type=mime,
            headers={"Content-Disposition": f'attachment; filename="{filename}"'}
        )
    except ValueError as e:
        raise HTTPException(status_code=400, detail=f"{e}")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"excel_error: {e}")
