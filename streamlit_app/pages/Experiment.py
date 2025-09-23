import streamlit as st
import yfinance as yf
import pandas as pd
import plotly.graph_objects as go
from datetime import datetime
import numpy as np
from scipy.stats import multivariate_normal
import io
from dateutil.relativedelta import relativedelta
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.cell import MergedCell
from openpyxl.utils import get_column_letter  # <-- avoids column_letter linter gripe

# ---------- Page & Style ----------
st.set_page_config(page_title="Quant Analytics Dashboard", layout="wide")

st.markdown(
    """
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap');
    .stApp { background: linear-gradient(135deg, #0f172a 0%, #1e293b 50%, #334155 100%); color: #f8fafc; font-family: 'Inter', -apple-system, BlinkMacSystemFont, sans-serif; }
    h1, h2, h3 { color: #22c55e; font-weight: 700; letter-spacing: -0.025em; }
    h1 { font-size: 2.8em; background: linear-gradient(135deg, #22c55e, #16a34a); -webkit-background-clip: text; -webkit-text-fill-color: transparent; }
    .section-container { background: rgba(15, 23, 42, 0.8); border: 1px solid rgba(34, 197, 94, 0.3); border-radius: 16px; padding: 30px; margin: 25px 0; box-shadow: 0 8px 32px rgba(0,0,0,0.4); backdrop-filter: blur(12px); }
    .highlight-box { background: rgba(15, 23, 42, 0.9); border: 1px solid rgba(34, 197, 94, 0.4); padding: 20px; border-radius: 12px; color: #f1f5f9 !important; }
    .highlight-box strong { color: #22c55e !important; }
    .stButton > button, .stDownloadButton button {
        background: linear-gradient(135deg, #22c55e 0%, #16a34a 100%) !important;
        color: #0f172a !important; border: none !important; border-radius: 10px !important; font-weight: 600 !important; padding: 12px 24px !important;
        transition: all 0.3s ease !important; box-shadow: 0 4px 12px rgba(34, 197, 94, 0.3);
    }
    .stButton > button:hover, .stDownloadButton button:hover { transform: translateY(-2px); box-shadow: 0 6px 20px rgba(34, 197, 94, 0.4); }
    .stTabs [data-baseweb="tab-list"] { gap: 15px; background: rgba(15, 23, 42, 0.6); border-radius: 12px; padding: 8px; }
    .stTabs [data-baseweb="tab"] { background: rgba(34, 197, 94, 0.1); color: #cbd5e1 !important; border-radius: 8px; padding: 12px 24px; font-weight: 600; border: 1px solid rgba(34, 197, 94, 0.2); }
    .stTabs [data-baseweb="tab"][aria-selected="true"] { background: linear-gradient(135deg, #22c55e, #16a34a) !important; color: #0f172a !important; box-shadow: 0 4px 12px rgba(34, 197, 94, 0.4); }
    .backtest-container { background: rgba(15, 23, 42, 0.9); border: 2px solid #22c55e; border-radius: 20px; padding: 32px; margin: 32px 0; box-shadow: 0 12px 40px rgba(34,197,94,0.3); text-align: center; }
    .metric-card { background: rgba(15, 23, 42, 0.8); border: 1px solid rgba(34, 197, 94, 0.4); padding: 16px; border-left: 4px solid #22c55e; font-weight: 600; }
    [data-testid="stSidebar"] * { color: #f1f5f9 !important; }
    </style>
    """,
    unsafe_allow_html=True
)

st.markdown(
    """
    <div style='text-align: center; padding: 30px; background: rgba(15, 23, 42, 0.8); border-radius: 16px; margin-bottom: 30px; border: 1px solid rgba(34, 197, 94, 0.3);'>
        <h1>Quantitative Analytics Dashboard</h1>
        <p style='color: #cbd5e1; font-size: 1.15em; max-width: 900px; margin: auto; line-height: 1.6;'>
            Select a pair, set a date range, explore 2D & 3D visuals, then review backtest results.
        </p>
    </div>
    """,
    unsafe_allow_html=True
)

# ---------- Sidebar (unique keys for multipage!) ----------
with st.sidebar:
    st.markdown(
        """
        <div style='text-align: center; padding: 16px; background: rgba(34, 197, 94, 0.1); border-radius: 12px; margin-bottom: 16px; border: 1px solid rgba(34, 197, 94, 0.3);'>
            <h2 style='margin: 0; color: #22c55e;'>Configuration</h2>
            <p style='margin: 8px 0 0 0; color: #cbd5e1; font-size: 0.95em;'>Pick parameters below</p>
        </div>
        """,
        unsafe_allow_html=True
    )

    pair_sel = st.selectbox(
        "Stock Pair",
        [
            ("NVDA", "AMD"), ("INTC", "AMD"), ("AVGO", "QCOM"), ("MU", "WDC"),
            ("GOOGL", "META"), ("MSFT", "ORCL"), ("V", "MA"), ("JPM", "BAC"),
            ("GS", "MS"), ("XOM", "CVX"), ("SLB", "HAL"), ("KO", "PEP"),
            ("WMT", "TGT"), ("HD", "LOW"), ("UPS", "FDX"), ("GM", "F"), ("DAL", "UAL")
        ],
        format_func=lambda t: f"{t[0]} vs {t[1]}",
        key="pair_exp_page",  # unique
    )

    start_dt = st.date_input("Start Date", value=datetime(2023, 1, 1), min_value=datetime(2015, 1, 1), key="start_exp_page")
    end_dt   = st.date_input("End Date",   value=datetime(2024, 1, 1), min_value=datetime(2015, 1, 1), key="end_exp_page")
    capital  = st.slider("Initial Capital ($)", 100, 5000, 1000, 100, key="cap_exp_page")

    st.markdown(
        """
        <div class='highlight-box'>
            <strong>Tips:</strong><br>
            • Use sector-matched pairs for cleaner signals<br>
            • Longer windows → more stable stats<br>
            • Watch rolling correlation dips
        </div>
        """,
        unsafe_allow_html=True
    )

if start_dt >= end_dt:
    st.error("End date must exceed start date.")
    st.stop()

# ---------- Data ----------
@st.cache_data(show_spinner=True)
def load_prices_cached(t1: str, t2: str, start, end) -> pd.DataFrame:
    tick_list = [t1, t2]
    df = yf.download(tick_list, start=start, end=end, progress=False)
    if df.empty:
        return pd.DataFrame()
    if ("Adj Close", t1) in df.columns:
        out = df.loc[:, ("Adj Close", slice(None))]
    else:
        out = df.loc[:, ("Close", slice(None))]
    out.columns = out.columns.droplevel(0)
    return out

sym_a, sym_b = pair_sel
price_df = load_prices_cached(sym_a, sym_b, start_dt, end_dt)
if price_df.empty:
    st.error("No data available for selected parameters.")
    st.stop()

st.success("Data successfully retrieved.")

# ---------- 2D ----------
st.markdown(
    """
    <div class='section-container'>
        <h2>2D Graphical Representations</h2>
        <p style='color:#cbd5e1;'>Price trajectories and the spread z-score.</p>
    </div>
    """,
    unsafe_allow_html=True
)
tab_price, tab_z = st.tabs(["Price Plot", "Z-Score Plot"])

with tab_price:
    fig_prices = go.Figure()
    fig_prices.add_trace(go.Scatter(x=price_df.index, y=price_df[sym_a], name=sym_a, mode="lines"))
    fig_prices.add_trace(go.Scatter(x=price_df.index, y=price_df[sym_b], name=sym_b, mode="lines"))
    fig_prices.update_layout(title=f"Price Trajectories: {sym_a} vs {sym_b}", xaxis_title="Date", yaxis_title="Price ($)", template="plotly_white")
    st.plotly_chart(fig_prices, use_container_width=True)

returns_df = price_df.pct_change().dropna()
spread_ser = returns_df[sym_a] - returns_df[sym_b]
z_ser = (spread_ser - spread_ser.mean()) / spread_ser.std(ddof=0)

with tab_z:
    fig_z = go.Figure()
    fig_z.add_trace(go.Scatter(x=z_ser.index, y=z_ser, name="Z-Score", mode="lines"))
    fig_z.add_hline(y=2, line_dash="dash", line_color="green", annotation_text="+2σ")
    fig_z.add_hline(y=-2, line_dash="dash", line_color="red", annotation_text="-2σ")
    fig_z.update_layout(title=f"Z-Score of Daily Return Spread: {sym_a} − {sym_b}", xaxis_title="Date", yaxis_title="Z-Score", template="plotly_white")
    st.plotly_chart(fig_z, use_container_width=True)

# ---------- 3D (robust) ----------
st.markdown(
    """
    <div class='section-container'>
        <h2>3D Analytical Visualizations</h2>
        <p style='color:#cbd5e1;'>Joint return distribution and rolling correlation surface.</p>
    </div>
    """,
    unsafe_allow_html=True
)
tab_joint, tab_roll = st.tabs(["123484949849", "Rolling Correlation Surface"])

# Joint distribution (regularize covariance to avoid singularities)
ra = (returns_df[sym_a].to_numpy() * 100.0)
rb = (returns_df[sym_b].to_numpy() * 100.0)
# guard against too-short or constant data
if len(ra) < 20 or np.allclose(np.nanstd(ra), 0) or np.allclose(np.nanstd(rb), 0):
    with tab_joint:
        st.warning("Not enough variation to plot a stable 3D distribution for this window.")
else:
    mean_vec = [float(np.nanmean(ra)), float(np.nanmean(rb))]
    cov_mat = np.cov(np.vstack([ra, rb]))
    # ensure PSD
    min_eig = np.min(np.real(np.linalg.eigvals(cov_mat)))
    if min_eig < 1e-8:
        cov_mat += (1e-8 - min_eig + 1e-8) * np.eye(2)

    dist = multivariate_normal(mean=mean_vec, cov=cov_mat)
    gx = np.linspace(np.nanmin(ra), np.nanmax(ra), 60)
    gy = np.linspace(np.nanmin(rb), np.nanmax(rb), 60)
    GX, GY = np.meshgrid(gx, gy)
    POS = np.dstack((GX, GY))
    GZ = dist.pdf(POS)
    GZ = np.nan_to_num(GZ, nan=0.0, posinf=0.0, neginf=0.0)

    with tab_joint:
        fig3d = go.Figure(data=[go.Surface(x=GX, y=GY, z=GZ, opacity=0.92)])
        fig3d.update_layout(
            title=f"Joint Distribution of Returns (%): {sym_a} vs {sym_b}",
            scene=dict(xaxis_title=f"{sym_a} Daily Return (%)", yaxis_title=f"{sym_b} Daily Return (%)", zaxis_title="Density"),
            width=900, height=700, template="plotly_white"
        )
        st.plotly_chart(fig3d, use_container_width=True)
        corr_val = float(np.corrcoef(ra, rb)[0, 1])
        st.markdown(f"<div class='highlight-box'><strong>Pearson Correlation: {corr_val:.2f}</strong></div>", unsafe_allow_html=True)

# Rolling correlation surface (windows x time), shapes must match
win_list = [30, 60, 90]
T = len(returns_df)
with tab_roll:
    if T < max(win_list) + 5:
        st.warning("Date range too short for rolling windows 30/60/90.")
    else:
        rc_rows = []
        for w in win_list:
            rc_rows.append(returns_df[sym_a].rolling(w).corr(returns_df[sym_b]).to_numpy())
        zmat = np.vstack(rc_rows)  # shape: (len(wins), T)
        t_axis = np.arange(zmat.shape[1])
        Tmesh, Wmesh = np.meshgrid(t_axis, np.array(win_list))
        # Replace crazy values
        zmat = np.nan_to_num(zmat, nan=np.nan, posinf=np.nan, neginf=np.nan)

        fig_roll = go.Figure(data=[go.Surface(x=Tmesh, y=Wmesh, z=zmat, opacity=0.9)])
        fig_roll.update_layout(
            title=f"Rolling Correlation Surface: {sym_a} vs {sym_b}",
            scene=dict(xaxis_title="Time Index", yaxis_title="Rolling Window (days)", zaxis_title="Correlation"),
            width=900, height=700, template="plotly_white"
        )
        st.plotly_chart(fig_roll, use_container_width=True)
        avgc = float(np.nanmean(zmat))
        st.markdown(f"<div class='highlight-box'><strong>Average rolling corr: {avgc:.2f}</strong></div>", unsafe_allow_html=True)

# ---------- Backtest ----------
rel = relativedelta(end_dt, start_dt)
period_str = (f"{rel.years} year{'s' if rel.years!=1 else ''}, {rel.months} month{'s' if rel.months!=1 else ''}"
              if rel.years and rel.months else
              (f"{rel.years} year{'s' if rel.years!=1 else ''}" if rel.years else f"{rel.months} month{'s' if rel.months!=1 else ''}"))

cum = (1 + returns_df).cumprod() - 1
lead = sym_a if float(cum[sym_a].iloc[-1]) > float(cum[sym_b].iloc[-1]) else sym_b
lagg = sym_b if lead == sym_a else sym_a
p_long  = capital * float(cum[lead].iloc[-1])
p_short = -capital * float(cum[lagg].iloc[-1])
p_net   = p_long + p_short
ann_vol = float(returns_df[[sym_a, sym_b]].std().mean() * np.sqrt(252))

fig_perf = go.Figure()
fig_perf.add_trace(go.Bar(
    x=[lead, lagg, "Net Profit"],
    y=[float(cum[lead].iloc[-1])*100, float(cum[lagg].iloc[-1])*100, p_net/capital*100],
    text=[f"{float(cum[lead].iloc[-1])*100:.2f}%", f"{float(cum[lagg].iloc[-1])*100:.2f}%", f"${p_net:.2f}"],
    textposition='auto'
))
fig_perf.update_layout(title="Performance Summary", yaxis_title="Return (%)", template="plotly_white", showlegend=False)

report_df = pd.DataFrame({
    "Metric": ["Stock Pair", "Outperforming Asset", "Underperforming Asset", "Net Profit", "Annualized Volatility", "Time Period"],
    "Value": [f"{sym_a} vs {sym_b}",
              f"{lead} ({float(cum[lead].iloc[-1])*100:.2f}%)",
              f"{lagg} ({float(cum[lagg].iloc[-1])*100:.2f}%)",
              f"Long ${capital} {lead}, Short ${capital} {lagg}: ${p_net:.2f}",
              f"{ann_vol*100:.2f}% (Risk Measure)",
              period_str]
})

def build_excel(df: pd.DataFrame, pa: str, pb: str) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Backtest Report"
    title = f"Pairs Trading Report ({pa} vs {pb})"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=df.shape[1])
    c = ws.cell(row=1, column=1, value=title)
    c.font = Font(bold=True, size=14, color="FFFFFF")
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.fill = PatternFill(start_color="22C55E", end_color="22C55E", fill_type="solid")

    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="22C55E", end_color="22C55E", fill_type="solid")
    thin = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

    for j, name in enumerate(df.columns, 1):
        cell = ws.cell(row=2, column=j, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin

    for r, rowvals in enumerate(df.values.tolist(), 3):
        for cidx, val in enumerate(rowvals, 1):
            cell = ws.cell(row=r, column=cidx, value=val)
            cell.alignment = Alignment(wrap_text=True, vertical="center")
            cell.border = thin
            cell.font = Font(size=11)

    # Safe column width sizing (no direct .column_letter on MergedCell)
    for col_idx in range(1, df.shape[1] + 1):
        letter = get_column_letter(col_idx)
        max_len = 0
        for r in range(1, 3 + len(df.values.tolist())):
            v = ws.cell(row=r, column=col_idx).value
            max_len = max(max_len, len(str(v)) if v is not None else 0)
        ws.column_dimensions[letter].width = max_len + 4

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()

st.markdown(
    f"""
    <div class='backtest-container'>
        <h2>Backtesting Results</h2>
        <div style='display:flex;flex-wrap:wrap;gap:16px;justify-content:center;'>
            <div class='metric-card'><p><strong>Outperformer</strong><br>{lead} ({float(cum[lead].iloc[-1])*100:.2f}%)</p></div>
            <div class='metric-card'><p><strong>Underperformer</strong><br>{lagg} ({float(cum[lagg].iloc[-1])*100:.2f}%)</p></div>
            <div class='metric-card'><p><strong>Net Profit</strong><br>Long ${capital} {lead}, Short ${capital} {lagg}: ${p_net:.2f}</p></div>
            <div class='metric-card'><p><strong>Ann. Volatility</strong><br>{ann_vol*100:.2f}%</p></div>
            <div class='metric-card'><p><strong>Period</strong><br>{period_str}</p></div>
        </div>
    </div>
    """,
    unsafe_allow_html=True
)

c1, c2 = st.columns([3, 1])
with c1:
    st.plotly_chart(fig_perf, use_container_width=True)
with c2:
    # Build bytes now & use a UNIQUE key (multipage-safe)
    excel_bytes = build_excel(report_df, sym_a, sym_b)
    st.write(f"File size: {len(excel_bytes)} bytes")  # quick sanity check
    st.download_button(
        label="Download Report",
        data=excel_bytes,
        file_name=f"backtest_report_{sym_a}_{sym_b}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key=f"dl_{sym_a}_{sym_b}_{start_dt}_{end_dt}"
    )

st.caption("Educational use only — not financial advice.")
